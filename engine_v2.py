"""
engine_v2.py — HubSpot sync backend for V2.
Supports user-defined column mapping, more contact/company properties,
and richer logging with a downloadable results table.
"""
import io
import time
import csv
import requests
import openpyxl
from datetime import datetime, timezone
from typing import Optional, Callable

BASE_URL = "https://api.hubapi.com"

# v3 association type IDs
ASSOC_CONTACT_COMPANY = 1
ASSOC_CONTACT_DEAL    = 4
ASSOC_COMPANY_DEAL    = 6

# Sheets to skip
NON_LEAD_KEYWORDS = ("funnel", "reporting", "leadgen", "basic", "pivot", "summary")

# ─── HubSpot field definitions ────────────────────────────────────────────────
# Each entry: (display_label, hubspot_property, required, field_type)
# field_type: "text" | "bool"

CONTACT_FIELDS = [
    ("Contact full name",    "full_name",                True,  "text"),
    ("Email address",        "email",                    True,  "text"),
    ("Job title",            "jobtitle",                 False, "text"),
    ("Phone number",         "phone",                    False, "text"),
    ("LinkedIn profile URL", "hs_linkedin_profile_url",  False, "text"),
]

COMPANY_FIELDS = [
    ("Company name",         "company_name",             True,  "text"),
    ("Company website",      "website",                  False, "text"),
]

STAGE_FIELDS = [
    ("Contacted (TRUE/FALSE)",         "contacted", True,  "bool"),
    ("Responded (TRUE/FALSE)",         "responded", True,  "bool"),
    ("Meeting scheduled (TRUE/FALSE)", "meeting",   True,  "bool"),
]

ALL_FIELDS = COMPANY_FIELDS + CONTACT_FIELDS + STAGE_FIELDS


# ─── Excel inspection ─────────────────────────────────────────────────────────

def get_excel_info(file_bytes: bytes) -> dict:
    """
    Returns info about the Excel file for the column-mapping UI:
      {
        "sheets": [sheet_name, ...],
        "all_headers": [header, ...],          # union of all headers across sheets
        "sample": {sheet: [[row values], ...]}, # first 3 data rows per sheet
      }
    All header values are unique strings (duplicates get sheet-prefix).
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    result = {"sheets": [], "all_headers": [], "sample": {}}
    seen_headers = {}

    for sheet_name in wb.sheetnames:
        if any(k in sheet_name.lower() for k in NON_LEAD_KEYWORDS):
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4:
            continue

        header_row = rows[2]  # row 3 = column headers
        data_rows  = rows[3:6]  # first 3 data rows for preview

        # Build header list for this sheet
        sheet_headers = []
        for i, cell in enumerate(header_row):
            label = str(cell).strip() if cell else f"Column {i+1}"
            # Make unique across sheets
            if label in seen_headers and seen_headers[label] != sheet_name:
                label = f"{label} ({sheet_name})"
            seen_headers[label] = sheet_name
            sheet_headers.append(label)

        result["sheets"].append(sheet_name)
        result["sample"][sheet_name] = [
            [str(v) if v is not None else "" for v in row[:len(sheet_headers)]]
            for row in data_rows
        ]

        # Merge into all_headers (keep unique, in order)
        for h in sheet_headers:
            if h not in result["all_headers"]:
                result["all_headers"].append(h)

    wb.close()
    return result


def _as_bool(v) -> bool:
    return v is True or v == 1 or (isinstance(v, str) and v.strip().upper() == "TRUE")


def _is_placeholder(s: str) -> bool:
    return s.lower().strip() in ("lusha", "none", "n/a", "na", "-", "")


# ─── Excel reader with user mapping ──────────────────────────────────────────

def read_excel_mapped(file_bytes: bytes, field_to_header: dict,
                      all_headers: list) -> tuple:
    """
    Reads Excel using the user-defined column mapping.

    field_to_header: { field_key: header_string_or_None }
    all_headers: the ordered list of header strings (used to find indices)

    Returns:
        companies  – { company_name: [contact_dict, ...] }
        sheets     – list of processed sheet names
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    companies: dict = {}
    sheets_done: list = []

    for sheet_name in wb.sheetnames:
        if any(k in sheet_name.lower() for k in NON_LEAD_KEYWORDS):
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4:
            continue

        header_row = rows[2]
        data_rows  = rows[3:]

        # Build header→index map for this sheet
        sheet_header_map = {}
        for i, cell in enumerate(header_row):
            label = str(cell).strip() if cell else f"Column {i+1}"
            sheet_header_map[label] = i

        # Also try the with-suffix version (from get_excel_info)
        for h in all_headers:
            base = h.split(f" ({sheet_name})")[0]
            if base in sheet_header_map and h not in sheet_header_map:
                sheet_header_map[h] = sheet_header_map[base]

        def col(field_key) -> Optional[int]:
            header = field_to_header.get(field_key)
            if not header or header == "__skip__":
                return None
            return sheet_header_map.get(header)

        company_col  = col("company_name")
        if company_col is None:
            continue  # can't identify companies without this

        sheets_done.append(sheet_name)

        for row in data_rows:
            if not row:
                continue

            def get(idx) -> str:
                if idx is None or len(row) <= idx:
                    return ""
                v = row[idx]
                return str(v).strip().replace("\n", " ") if v is not None else ""

            company = get(company_col)
            if not company:
                continue

            # Email
            email_col = col("email")
            email_raw = get(email_col) if email_col is not None else ""
            email = None if _is_placeholder(email_raw) else email_raw or None

            # Name → split into first/last
            name_col = col("full_name")
            full_name = get(name_col) if name_col is not None else ""
            parts = full_name.split(None, 1)
            firstname = parts[0] if parts else ""
            lastname  = parts[1] if len(parts) > 1 else ""

            # Optional text fields
            def opt(field_key) -> Optional[str]:
                c = col(field_key)
                if c is None:
                    return None
                v = get(c)
                return None if _is_placeholder(v) else v or None

            # Stage booleans
            def stage_bool(field_key) -> bool:
                c = col(field_key)
                if c is None or len(row) <= c:
                    return False
                return _as_bool(row[c])

            companies.setdefault(company, []).append({
                "company":   company,
                "firstname": firstname,
                "lastname":  lastname,
                "email":     email,
                "jobtitle":  opt("jobtitle"),
                "phone":     opt("phone"),
                "linkedin":  opt("hs_linkedin_profile_url"),
                "website":   opt("website"),
                "industry":  sheet_name,
                "contacted": stage_bool("contacted"),
                "responded": stage_bool("responded"),
                "meeting":   stage_bool("meeting"),
            })

    wb.close()
    return companies, sheets_done


def determine_stage(contacts: list) -> Optional[str]:
    if any(c["meeting"]   for c in contacts): return "meeting"
    if any(c["responded"] for c in contacts): return "responded"
    if any(c["contacted"] for c in contacts): return "contacted"
    return None


def count_active(companies: dict) -> tuple:
    active = {c: v for c, v in companies.items() if determine_stage(v) is not None}
    emails = sum(1 for cs in active.values() for c in cs if c["email"])
    return len(companies), len(active), emails


# ─── HubSpot client ───────────────────────────────────────────────────────────

class HubSpotV2:
    def __init__(self, token: str):
        self.session = requests.Session()
        self.session.headers.update({
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        })
        self._last = 0.0

    def _wait(self):
        gap = time.time() - self._last
        if gap < 0.13:
            time.sleep(0.13 - gap)
        self._last = time.time()

    def _get(self, path):
        self._wait()
        r = self.session.get(f"{BASE_URL}{path}")
        r.raise_for_status()
        return r.json()

    def _post(self, path, body):
        self._wait()
        r = self.session.post(f"{BASE_URL}{path}", json=body)
        r.raise_for_status()
        return r.json()

    def _patch(self, path, body):
        self._wait()
        r = self.session.patch(f"{BASE_URL}{path}", json=body)
        r.raise_for_status()
        return r.json()

    def _put(self, path):
        self._wait()
        r = self.session.put(f"{BASE_URL}{path}")
        r.raise_for_status()

    def get_pipelines(self):
        return self._get("/crm/v3/pipelines/deals").get("results", [])

    def get_stages(self, pipeline_id):
        return self._get(f"/crm/v3/pipelines/deals/{pipeline_id}").get("stages", [])

    # ── Companies ──────────────────────────────────────────────────────────────

    def find_company(self, name) -> Optional[str]:
        data = self._post("/crm/v3/objects/companies/search", {
            "filterGroups": [{"filters": [
                {"propertyName": "name", "operator": "EQ", "value": name}
            ]}],
            "properties": ["name"], "limit": 1,
        })
        res = data.get("results", [])
        return res[0]["id"] if res else None

    def create_company(self, name, website=None) -> str:
        props = {"name": name}
        if website:
            props["website"] = website
        return self._post("/crm/v3/objects/companies", {"properties": props})["id"]

    def update_company(self, cid, website=None):
        props = {}
        if website:
            props["website"] = website
        if props:
            self._patch(f"/crm/v3/objects/companies/{cid}", {"properties": props})

    def get_or_create_company(self, name, website=None) -> tuple:
        cid = self.find_company(name)
        if cid:
            self.update_company(cid, website)
            return cid, False
        return self.create_company(name, website), True

    # ── Contacts ───────────────────────────────────────────────────────────────

    def find_contact(self, email) -> Optional[str]:
        data = self._post("/crm/v3/objects/contacts/search", {
            "filterGroups": [{"filters": [
                {"propertyName": "email", "operator": "EQ", "value": email}
            ]}],
            "properties": ["email"], "limit": 1,
        })
        res = data.get("results", [])
        return res[0]["id"] if res else None

    def create_contact(self, contact: dict) -> str:
        props = {
            "email":     contact["email"],
            "firstname": contact["firstname"],
            "lastname":  contact["lastname"],
        }
        for key in ("jobtitle", "phone", "hs_linkedin_profile_url"):
            src = "linkedin" if key == "hs_linkedin_profile_url" else key
            if contact.get(src):
                props[key] = contact[src]
        return self._post("/crm/v3/objects/contacts", {"properties": props})["id"]

    def update_contact(self, cid: str, contact: dict):
        """Update existing contact with any new non-empty properties."""
        props = {}
        if contact.get("firstname"): props["firstname"] = contact["firstname"]
        if contact.get("lastname"):  props["lastname"]  = contact["lastname"]
        for key in ("jobtitle", "phone"):
            if contact.get(key): props[key] = contact[key]
        if contact.get("linkedin"): props["hs_linkedin_profile_url"] = contact["linkedin"]
        if props:
            self._patch(f"/crm/v3/objects/contacts/{cid}", {"properties": props})

    def get_or_create_contact(self, contact: dict) -> tuple:
        if not contact.get("email"):
            return None, None
        cid = self.find_contact(contact["email"])
        if cid:
            self.update_contact(cid, contact)
            return cid, False
        return self.create_contact(contact), True

    # ── Deals ──────────────────────────────────────────────────────────────────

    def get_company_deals(self, co_id) -> list:
        data = self._get(f"/crm/v3/objects/companies/{co_id}/associations/deals")
        return [x["id"] for x in data.get("results", [])]

    def create_deal(self, name, pipeline_id, stage_id) -> str:
        return self._post("/crm/v3/objects/deals", {"properties": {
            "dealname": name, "pipeline": pipeline_id, "dealstage": stage_id,
        }})["id"]

    def update_deal_stage(self, deal_id, stage_id):
        self._patch(f"/crm/v3/objects/deals/{deal_id}",
                    {"properties": {"dealstage": stage_id}})

    # ── Associations ───────────────────────────────────────────────────────────

    def _link(self, from_t, from_id, to_t, to_id, type_id):
        try:
            self._put(f"/crm/v3/objects/{from_t}/{from_id}"
                      f"/associations/{to_t}/{to_id}/{type_id}")
        except requests.HTTPError as e:
            if e.response is not None and e.response.status_code == 409:
                return
            raise

    def link_contact_company(self, ct, co): self._link("contacts", ct, "companies", co, ASSOC_CONTACT_COMPANY)
    def link_contact_deal(self, ct, d):    self._link("contacts", ct, "deals", d, ASSOC_CONTACT_DEAL)
    def link_company_deal(self, co, d):    self._link("companies", co, "deals", d, ASSOC_COMPANY_DEAL)


# ─── Sync ─────────────────────────────────────────────────────────────────────

def run_sync_v2(companies: dict, hs: HubSpotV2,
                pipeline_id: str, stage_map: dict,
                log: Callable = print,
                fallback_stage: Optional[str] = None) -> tuple:
    """
    Returns (stats_dict, log_rows) where log_rows is a list of dicts
    for the downloadable CSV log.
    """
    stats = {
        "skipped": 0, "co_created": 0, "co_updated": 0,
        "ct_created": 0, "ct_updated": 0,
        "deal_created": 0, "deal_updated": 0, "errors": 0,
    }
    log_rows = []

    def record(company, name, email, action, stage, note=""):
        log_rows.append({
            "Timestamp": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
            "Company":   company,
            "Contact":   name,
            "Email":     email or "",
            "Action":    action,
            "Stage":     stage or "",
            "Note":      note,
        })

    for company_name, contacts in companies.items():
        stage = determine_stage(contacts)
        if stage is None:
            if fallback_stage is None:
                stats["skipped"] += 1
                record(company_name, "", "", "Skipped", None, "No activity")
                continue
            stage_id = fallback_stage  # already a HubSpot stage ID
            stage    = "imported"      # label for logs only
        else:
            stage_id = stage_map[stage]

        log(f"▸  {company_name}  →  {stage.upper()}")

        # Website from first contact that has it
        website = next((c["website"] for c in contacts if c.get("website")), None)

        # Company
        try:
            co_id, created = hs.get_or_create_company(company_name, website)
            if created:
                stats["co_created"] += 1
                log(f"   + company created")
                record(company_name, "", "", "Company Created", stage)
            else:
                stats["co_updated"] += 1
                log(f"   = company found & updated")
                record(company_name, "", "", "Company Updated", stage)
        except Exception as e:
            if isinstance(e, requests.HTTPError) and e.response is not None:
                code = e.response.status_code
                try:
                    body = e.response.json().get("message", e.response.text[:120])
                except Exception:
                    body = e.response.text[:120]
                detail = f"HTTP {code}: {body}"
            else:
                detail = f"{type(e).__name__}: {e}"
            log(f"   ! company error  {detail}")
            stats["errors"] += 1
            record(company_name, "", "", "Error", stage, f"Company {detail}")
            continue

        # Contacts
        ct_ids = []
        for c in contacts:
            name_str = f"{c['firstname']} {c['lastname']}".strip()
            if not c["email"]:
                log(f"   – {name_str}  (no email, skipped)")
                continue
            try:
                ct_id, created = hs.get_or_create_contact(c)
                if ct_id is None:
                    continue
                label = "Contact Created" if created else "Contact Updated"
                if created:
                    stats["ct_created"] += 1
                    log(f"   + contact {name_str} <{c['email']}>")
                else:
                    stats["ct_updated"] += 1
                    log(f"   = contact {name_str} <{c['email']}> updated")
                record(company_name, name_str, c["email"], label, stage)
                ct_ids.append(ct_id)
                hs.link_contact_company(ct_id, co_id)
            except Exception as e:
                if isinstance(e, requests.HTTPError) and e.response is not None:
                    code = e.response.status_code
                    try:
                        body = e.response.json().get("message", e.response.text[:120])
                    except Exception:
                        body = e.response.text[:120]
                    detail = f"HTTP {code}: {body}"
                else:
                    detail = f"{type(e).__name__}: {e}"
                log(f"   ! contact error {c['email']}  {detail}")
                stats["errors"] += 1
                record(company_name, name_str, c["email"], "Error", stage, detail)

        # Deal
        try:
            existing = hs.get_company_deals(co_id)
        except Exception:
            existing = []

        try:
            if existing:
                deal_id = existing[0]
                hs.update_deal_stage(deal_id, stage_id)
                stats["deal_updated"] += 1
                log(f"   = deal stage → {stage}")
            else:
                deal_id = hs.create_deal(company_name, pipeline_id, stage_id)
                stats["deal_created"] += 1
                log(f"   + deal created  (stage={stage})")

            hs.link_company_deal(co_id, deal_id)
            for ct_id in ct_ids:
                hs.link_contact_deal(ct_id, deal_id)

        except requests.HTTPError as e:
            code = e.response.status_code if e.response else "?"
            log(f"   ! deal error  HTTP {code}")
            stats["errors"] += 1
            record(company_name, "", "", "Deal Error", stage, f"HTTP {code}")

    return stats, log_rows


def log_rows_to_csv(log_rows: list) -> str:
    """Convert log rows to CSV string for download."""
    if not log_rows:
        return ""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=log_rows[0].keys())
    writer.writeheader()
    writer.writerows(log_rows)
    return buf.getvalue()
